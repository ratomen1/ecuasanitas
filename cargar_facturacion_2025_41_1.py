#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para cargar el archivo Excel FACTURACION 2025 41 Ingreso Operativos-1.xlsx
a la tabla FACTURACION_2025_41_1 en PostgreSQL 9.5.

Uso:
    python3 cargar_facturacion_2025_41_1.py

El script crea un virtualenv automáticamente e instala las dependencias.
"""

import subprocess
import sys
import os

# ============================================================
# AUTO-SETUP: Crear virtualenv e instalar dependencias
# ============================================================
VENV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.venv')
VENV_PYTHON = os.path.join(VENV_DIR, 'bin', 'python3')

def setup_venv():
    """Crea el virtualenv e instala dependencias si no existen."""
    if not os.path.exists(VENV_PYTHON):
        print("Creando virtualenv en .venv/ ...")
        subprocess.check_call([sys.executable, '-m', 'venv', VENV_DIR])
        print("Instalando dependencias (openpyxl, psycopg2-binary)...")
        subprocess.check_call([
            os.path.join(VENV_DIR, 'bin', 'pip'),
            'install', '--quiet', 'openpyxl', 'psycopg2-binary'
        ])
        print("Dependencias instaladas.\n")

    # Re-ejecutar este mismo script con el Python del venv
    if os.path.abspath(sys.executable) != os.path.abspath(VENV_PYTHON):
        os.execv(VENV_PYTHON, [VENV_PYTHON] + sys.argv)

setup_venv()

# A partir de aquí ya estamos corriendo dentro del venv
import openpyxl
import psycopg2
import re
import io
import math
from datetime import datetime, date

# ============================================================
# CONFIGURACIÓN — Ajustar según el entorno
# ============================================================
EXCEL_PATH = '/home/jose/Descargas/caidos 2023_2026-02.xlsx'
TABLE_NAME = 'caidos_2023_2026_02'
SQL_OUTPUT = '/home/jose/Documentos/sandbox/ecuasanitas/PRODUCCION_GENESYS/cargar_facturacion_2025_41_1.sql'

DB_CONFIG = {
    'dbname': 'genesys',
    'user': 'genesys',
    'password': 'genesys',
    'host': '192.168.40.128',
    'port': '5432'
}

# ============================================================
# FUNCIONES
# ============================================================

def sanitize_column_name(name):
    """Convierte el nombre de columna del Excel a un nombre válido para PostgreSQL."""
    if name is None:
        return 'columna_sin_nombre'
    col = str(name).strip().lower()
    col = col.replace(' ', '_').replace('.', '_').replace('-', '_')
    col = col.replace('á', 'a').replace('é', 'e').replace('í', 'i')
    col = col.replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
    col = re.sub(r'[^a-z0-9_]', '', col)
    col = re.sub(r'_+', '_', col).strip('_')
    if not col:
        col = 'col'
    if col[0].isdigit():
        col = '_' + col
    return col


def value_to_str(v):
    """Convierte un valor a string preservando decimales exactos.
    Retorna None si el valor es nulo/vacío/inválido."""
    if v is None:
        return None
    try:
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, float):
            if math.isnan(v) or math.isinf(v):
                return None
            # Preservar decimales exactos: usar f-string con suficiente precisión
            # y eliminar ceros trailing innecesarios
            formatted = f'{v:.15g}'
            return formatted
        if isinstance(v, int):
            return str(v)
        if isinstance(v, datetime):
            return v.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(v, date):
            return v.strftime('%Y-%m-%d')
        if isinstance(v, str):
            stripped = v.strip()
            return stripped if stripped else None
        return str(v)
    except Exception:
        return None


def read_excel(path):
    """Lee el Excel y retorna headers y filas de datos (todo como strings)."""
    print(f"Leyendo archivo: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    print(f"  Hoja: {ws.title}")
    print(f"  Filas: {ws.max_row}, Columnas: {ws.max_column}")

    rows_iter = ws.iter_rows()
    header_row = next(rows_iter)
    headers = [sanitize_column_name(cell.value) for cell in header_row]

    # Detectar duplicados en headers
    seen = {}
    for i, h in enumerate(headers):
        if h in seen:
            seen[h] += 1
            headers[i] = f"{h}_{seen[h]}"
        else:
            seen[h] = 0

    # Leer todos los datos, convertir a strings
    data = []
    total = 0
    for row in rows_iter:
        raw_vals = [cell.value for cell in row]
        # Saltar filas completamente vacías
        if not any(v is not None for v in raw_vals):
            continue

        total += 1
        str_vals = [value_to_str(v) for v in raw_vals]

        # Construir observacion: detectar NULLs
        errores = []
        for col_idx, (raw, clean) in enumerate(zip(raw_vals, str_vals)):
            col_name = headers[col_idx] if col_idx < len(headers) else f'col_{col_idx}'
            if clean is None:
                if raw is not None and str(raw).strip() != '':
                    errores.append(f"Error en la columna {col_name} el valor {repr(raw)} es null")
                else:
                    errores.append(f"Error en la columna {col_name} el valor es null")

        observacion = ' | '.join(errores) if errores else None
        str_vals.append(observacion)

        # Asegurar cantidad correcta de columnas (+1 por observacion)
        expected = len(headers) + 1
        while len(str_vals) < expected:
            str_vals.append(None)
        str_vals = str_vals[:expected]

        data.append(str_vals)

        if total % 100000 == 0:
            print(f"    {total} filas leídas...")

    # Agregar columna observacion
    headers.append('observacion')

    print(f"  Total filas de datos: {len(data)}")
    wb.close()
    return headers, data


def generate_create_table(table_name, headers):
    """Genera el DDL CREATE TABLE — todas las columnas como TEXT."""
    lines = [f'DROP TABLE IF EXISTS {table_name};']
    lines.append(f'CREATE TABLE {table_name} (')
    col_defs = [f'    {h} TEXT' for h in headers]
    lines.append(',\n'.join(col_defs))
    lines.append(');')
    return '\n'.join(lines)


def escape_copy_value(v):
    """Escapa un valor para COPY FROM (formato TSV).
    NULL -> \\N, tabs/newlines escapados."""
    if v is None:
        return '\\N'
    s = str(v)
    s = s.replace('\\', '\\\\')
    s = s.replace('\t', '\\t')
    s = s.replace('\n', '\\n')
    s = s.replace('\r', '\\r')
    return s


def load_to_database(db_config, table_name, headers, data):
    """Carga los datos directamente a PostgreSQL usando COPY (ultra rápido)."""
    print("\nConectando a la base de datos...")
    conn = None
    try:
        conn = psycopg2.connect(**db_config)
        conn.autocommit = False
        cur = conn.cursor()

        # Crear tabla (todas TEXT)
        create_ddl = generate_create_table(table_name, headers)
        print(f"  Creando tabla {table_name} (todas las columnas TEXT)...")
        cur.execute(f"DROP TABLE IF EXISTS {table_name}")
        cur.execute(create_ddl.replace(f"DROP TABLE IF EXISTS {table_name};", "").strip())
        conn.commit()

        # Construir buffer TSV en memoria para COPY
        print(f"  Preparando {len(data)} filas para COPY...")
        buffer = io.StringIO()
        for row in data:
            line = '\t'.join(escape_copy_value(v) for v in row)
            buffer.write(line + '\n')
        buffer.seek(0)

        # COPY FROM stdin — carga masiva ultra rápida
        cols = ', '.join(headers)
        copy_sql = f"COPY {table_name} ({cols}) FROM STDIN WITH (FORMAT text, DELIMITER E'\\t', NULL '\\N')"
        print(f"  Ejecutando COPY...")
        cur.copy_expert(copy_sql, buffer)

        # Verificar
        cur.execute(f"SELECT COUNT(*) FROM {table_name}")
        count = cur.fetchone()[0]
        print(f"  Total registros en tabla: {count}")

        conn.commit()
        print("  COMMIT exitoso.")

        cur.close()
        conn.close()
        print("  Conexión cerrada.")
        return True

    except Exception as e:
        print(f"\n  ERROR al cargar en BD: {e}")
        print("  Se hizo ROLLBACK. Los datos NO se insertaron.")
        if conn:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
        return False


def write_sql_file(path, table_name, headers, data):
    """Escribe el archivo SQL completo como respaldo."""
    print(f"  Generando archivo SQL de respaldo...")
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f"-- ============================================================\n")
        f.write(f"-- Carga de datos: {table_name}\n")
        f.write(f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- Fuente: FACTURACION 2025 41 Ingreso Operativos-1.xlsx\n")
        f.write(f"-- Total filas: {len(data)}\n")
        f.write(f"-- TODAS las columnas son TEXT para evitar errores de tipo\n")
        f.write(f"-- ============================================================\n\n")
        f.write("BEGIN;\n\n")
        f.write(generate_create_table(table_name, headers))
        f.write('\n\n')

        # Inserts en lotes de 100
        cols = ', '.join(headers)
        batch_size = 100
        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            values_list = []
            for row in batch:
                vals = []
                for v in row:
                    if v is None:
                        vals.append('NULL')
                    else:
                        escaped = str(v).replace("'", "''")
                        vals.append(f"'{escaped}'")
                values_list.append(f"    ({', '.join(vals)})")
            f.write(f"INSERT INTO {table_name} ({cols})\nVALUES\n")
            f.write(',\n'.join(values_list) + ';\n\n')

        f.write(f"SELECT COUNT(*) AS total_registros FROM {table_name};\n\n")
        f.write("COMMIT;\n")
    print(f"  Archivo SQL generado: {path}")


# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    # 1. Leer Excel
    headers, data = read_excel(EXCEL_PATH)

    print(f"\nColumnas detectadas ({len(headers)}):")
    for h in headers:
        print(f"  {h:40s} TEXT")

    # Contar filas con observaciones
    obs_idx = headers.index('observacion')
    filas_con_error = sum(1 for row in data if row[obs_idx] is not None)
    print(f"\n  Filas con valores NULL detectados: {filas_con_error}")
    print(f"  Filas limpias: {len(data) - filas_con_error}")

    # 2. Guardar archivo SQL de respaldo
    write_sql_file(SQL_OUTPUT, TABLE_NAME, headers, data)

    # 3. Preguntar si cargar directo a BD
    print("\n" + "=" * 60)
    print("¿Desea cargar los datos directamente a la base de datos?")
    print(f"  Host: {DB_CONFIG['host']}:{DB_CONFIG['port']}")
    print(f"  BD:   {DB_CONFIG['dbname']}")
    print(f"  Tabla: {TABLE_NAME}")
    print(f"  Columnas: TODAS como TEXT (sin errores de tipo)")
    print(f"  Método: COPY (carga masiva ultra rápida)")
    print("=" * 60)
    resp = input("\nEscriba 'si' para cargar, cualquier otra cosa para solo generar SQL: ").strip().lower()

    if resp in ('si', 'sí', 's', 'yes', 'y'):
        load_to_database(DB_CONFIG, TABLE_NAME, headers, data)
    else:
        print("\nSolo se generó el archivo SQL. Puede ejecutarlo manualmente en DataGrip.")
        print(f"  Archivo: {SQL_OUTPUT}")

    print("\n¡Listo!")
